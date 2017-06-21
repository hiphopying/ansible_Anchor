# This script load LVSW excel and read all OS IPs and write them to a file

import re
from openpyxl import workbook, worksheet, load_workbook


WNH = {'ip':10} #column index for owner, IP, and mangement for each sheet
VMH = {'ip':10}
SOH = {'ip':10}
LXH = {'ip':10}
SGH = {'ip':10}
SCALEIO = {'ip':10}
DSSD = {'ip':9}
UCS_blades = {'ip':10}
#sheetname mapping to sheet index for LVSW xlsx
lvsw_sheets = {'SGH':SGH, 'WNH':WNH,'VMH':VMH,'SOH':SOH,'LXH':LXH, 'UCS Chassis & '\
               'Blades':UCS_blades, 'ScaleIO 12AQ':SCALEIO, 'DSSD 12AR':DSSD} 

def load_LVSW():
    return load_workbook('LVSW_host_list.xlsx')

def match_ip(ip):
    if re.match(r"(?P<IP>^\d*\.\d*\.\d*\.\d*$)",ip):
        return 1
    else:
        return 0


def write_ips_tofile(path):
    wb = load_LVSW()
    #iterate every sheet and rows
    with open(path, 'w') as f:
        f.write("reserved_m_ip:\n")
        for sheet in lvsw_sheets.iterkeys():
            ws = wb.get_sheet_by_name(sheet)
            for row in ws.iter_rows():
                ip = str(row[lvsw_sheets.get(sheet).get('ip')].value) #check column "IP" for this row
                if match_ip(ip):
                    f.write("  - "+ip+"\n")

write_ips_tofile("./roles/render/vars/ip_inventory.yml")
