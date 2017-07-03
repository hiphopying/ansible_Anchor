##!/usr/bin/python
# -*- coding: utf-8 -*-

# This file is part of Ansible
#
# Ansible is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Ansible is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Ansible.  If not, see <http://www.gnu.org/licenses/>.

ANSIBLE_METADATA = {'metadata_version': '1.0',
                    'status': ['preview'],
                    'supported_by': 'community'}

DOCUMENTATION = '''

'''

EXAMPLES = '''

'''

# This script load LVSW excel and read all OS IPs and write them to a file

import re
from openpyxl import workbook, worksheet, load_workbook

LVSWPATH = "LVSW_host_list.xlsx"
OUTPUTFILE = "./roles/render/vars/ip_inventory.yml"
WNH = {'ip':10} #column index for owner, IP, and mangement for each sheet
VMH = {'ip':10}
SOH = {'ip':10}
LXH = {'ip':10}
SGH = {'ip':10}
SCALEIO = {'ip':10}
DSSD = {'ip':12}
UCS_blades = {'ip':10}
#sheetname mapping to sheet index for LVSW xlsx
lvsw_sheets = {'SGH':SGH, 'WNH':WNH,'VMH':VMH,'SOH':SOH,'LXH':LXH, 'UCS Chassis & '\
               'Blades':UCS_blades, 'ScaleIO 12AQ':SCALEIO, 'DSSD 12AR':DSSD} 

def load_LVSW(lvswpath):
    return load_workbook(LVSWPATH)

def match_ip(ip):
    if re.match(r"(?P<IP>^\d*\.\d*\.\d*\.\d*$)",ip.strip()):
        return 1
    else:
        return 0


def main():
    module = AnsibleModule(
        argument_spec = dict(
            lvswfile=dict(default=LVSWPATH, type='path'),
            outputfile=dict(default=OUTPUTFILE, type='path')
        ),
        supports_check_mode = True
    )
    lvswfile = module.params['lvswfile']
    outputfile = module.params['outputfile']
    wb = load_LVSW(lvswfile)
    with open(outputfile, 'w') as f:
        if not f:
            module.exit_json(changed=False)
            
        f.write("reserved_m_ip:\n")
        for sheet in lvsw_sheets.iterkeys():
            ws = wb.get_sheet_by_name(sheet)
            for row in ws.iter_rows():
                ip = str(row[lvsw_sheets.get(sheet).get('ip')].value) #check column "IP" for this row
                if match_ip(ip):
                    f.write("  - "+ip+"\n")
    
    module.exit_json(changed=True)



from ansible.module_utils.basic import *
from ansible.module_utils.pycompat24 import get_exception

if __name__ == '__main__':
     main()
